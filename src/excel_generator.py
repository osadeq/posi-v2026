"""
Module de génération Excel pour les programmes de formation personnalisés
eS@deq - Export Excel avec planning, suivi, budget et ressources
"""
import json
from pathlib import Path
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.styles.protection import Protection
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation


def charger_programme(id_candidat, base_dir):
    """Charge le programme personnalisé d'un candidat"""
    prog_path = Path(base_dir) / "data" / "candidats" / id_candidat / "programme_perso.json"
    if not prog_path.exists():
        return None
    with open(prog_path, 'r', encoding='utf-8') as f:
        return json.load(f)


def charger_questions(base_dir):
    """Charge la base de questions pour les liens vers fiches"""
    q_path = Path(base_dir) / "data" / "bd" / "00_questions.json"
    with open(q_path, 'r', encoding='utf-8') as f:
        return json.load(f)


def creer_styles(wb):
    """Crée les styles personnalisés pour le classeur"""
    # Style titre
    titre = NamedStyle(name="titre")
    titre.font = Font(name='Calibri', size=16, bold=True, color="1F4E78")
    titre.alignment = Alignment(horizontal='center', vertical='center')
    wb.add_named_style(titre)
    
    # Style sous-titre
    sous_titre = NamedStyle(name="sous_titre")
    sous_titre.font = Font(name='Calibri', size=12, bold=True, color="2E75B6")
    sous_titre.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    sous_titre.alignment = Alignment(horizontal='left', vertical='center')
    wb.add_named_style(sous_titre)
    
    # Style en-tête
    entete = NamedStyle(name="entete")
    entete.font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    entete.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    entete.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    wb.add_named_style(entete)
    
    # Style cellule
    cellule = NamedStyle(name="cellule")
    cellule.font = Font(name='Calibri', size=10)
    cellule.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    wb.add_named_style(cellule)
    
    # Style case à cocher
    case_cocher = NamedStyle(name="case_cocher")
    case_cocher.font = Font(name='Wingdings', size=12)
    case_cocher.alignment = Alignment(horizontal='center', vertical='center')
    wb.add_named_style(case_cocher)
    
    # Style total
    total = NamedStyle(name="total")
    total.font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    total.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    total.alignment = Alignment(horizontal='right', vertical='center')
    wb.add_named_style(total)
    
    # Style besoin fort
    besoin_fort = NamedStyle(name="besoin_fort")
    besoin_fort.font = Font(name='Calibri', size=10, color="C00000", bold=True)
    besoin_fort.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    wb.add_named_style(besoin_fort)
    
    # Style besoin moyen
    besoin_moyen = NamedStyle(name="besoin_moyen")
    besoin_moyen.font = Font(name='Calibri', size=10, color="9C5700")
    besoin_moyen.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    wb.add_named_style(besoin_moyen)
    
    # Style lien
    lien = NamedStyle(name="lien")
    lien.font = Font(name='Calibri', size=10, color="0563C1", underline="single")
    wb.add_named_style(lien)


def creer_onglet_planning(ws, programme, candidat):
    """Crée l'onglet Planning détaillé"""
    # Titre
    ws.merge_cells('A1:H1')
    ws['A1'] = f"PLANNING DE FORMATION - {candidat.get('prenom', '')} {candidat.get('nom', '')}"
    ws['A1'].style = 'titre'
    ws.row_dimensions[1].height = 30
    
    # Info programme
    ws['A3'] = "Programme :"
    ws['B3'] = programme.get('programme_nom', '')
    ws['A4'] = "Date de génération :"
    ws['B4'] = datetime.now().strftime('%d/%m/%Y')
    ws['A5'] = "Durée totale prévue :"
    ws['B5'] = f"{programme.get('estimation', {}).get('total_heures', 0):.2f} heures".replace('.', ',')
    
    for row in [3, 4, 5]:
        ws[f'A{row}'].font = Font(bold=True)
    
    # En-têtes tableau
    headers = ['Séance', 'Date', 'Horaire', 'Thème(s)', 'Activités', 'Durée', 'Statut', 'Commentaires']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col, value=header)
        cell.style = 'entete'
    
    ws.row_dimensions[7].height = 25
    
    # Données du déroulé
    row = 8
    deroule = programme.get('deroule_pedagogique', {})
    
    for seance in deroule.get('seances', []):
        # Date calculée (à partir d'aujourd'hui + décalage)
        date_seance = (datetime.now() + timedelta(days=(seance['seance_numero']-1)*7)).strftime('%d/%m/%Y')
        
        # Horaire
        horaire = "09:00-12:30" if seance['seance_numero'] % 2 == 1 else "14:00-17:30"
        
        # Thèmes
        themes = ", ".join([t['nom'] for t in seance.get('themes_couverts', [])])
        
        # Activités
        activites = "\n".join([f"• {a['activite']}" for a in seance.get('activites', [])])
        
        # Durée
        duree = f"{seance['duree_heures']}h"
        
        data = [
            seance['seance_numero'],
            date_seance,
            horaire,
            themes,
            activites,
            duree,
            "☐",  # Case à cocher
            ""
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.style = 'cellule'
            
            # Style spécial pour la case à cocher
            if col == 7:
                cell.style = 'case_cocher'
                cell.font = Font(name='Wingdings', size=14)
        
        # Hauteur de ligne adaptative
        nb_activites = len(seance.get('activites', []))
        ws.row_dimensions[row].height = max(30, nb_activites * 15)
        
        row += 1
    
    # Certification si présente
    if deroule.get('certification'):
        cert = deroule['certification']
        ws.cell(row=row, column=1, value="CERTIFICATION")
        ws.cell(row=row, column=2, value=cert.get('type', 'TOSA'))
        ws.cell(row=row, column=4, value=cert.get('note', ''))
        ws.cell(row=row, column=6, value=cert.get('duree', '1h00'))
        ws.cell(row=row, column=7, value="☐")
        ws.cell(row=row, column=7).style = 'case_cocher'
        ws.cell(row=row, column=7).font = Font(name='Wingdings', size=14)
        row += 1
    
    # Ajustement des colonnes
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 30
    
    # Mise en page
    ws.print_title_rows = '7:7'
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToHeight = 1


def creer_onglet_suivi(ws, programme):
    """Crée l'onglet Suivi de progression avec cases à cocher"""
    # Titre
    ws.merge_cells('A1:F1')
    ws['A1'] = "SUIVI DE PROGRESSION"
    ws['A1'].style = 'titre'
    ws.row_dimensions[1].height = 30
    
    # En-têtes
    headers = ['Thème', 'Niveau besoin', 'Durée', 'Acquis ☐', 'En cours ☐', 'À revoir ☐']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.style = 'entete'
    
    ws.row_dimensions[3].height = 25
    
    # Données
    row = 4
    theme_index = 0
    for theme in programme.get('themes_a_former', []):
        theme_index += 1
        niveau = theme.get('niveau_besoin', '')

        ws.cell(row=row, column=1, value=f"Thème {theme_index} : {theme['nom']}")
        ws.cell(row=row, column=2, value=niveau)
        ws.cell(row=row, column=3, value=f"{float(theme['duree_estimee']):.2f}".replace('.', ','))
        
        # Cases à cocher pour le suivi
        ws.cell(row=row, column=4, value="☐")
        ws.cell(row=row, column=5, value="☐")
        ws.cell(row=row, column=6, value="☐")
        
        # Styles
        ws.cell(row=row, column=1).style = 'cellule'
        
        if 'fort' in niveau.lower():
            ws.cell(row=row, column=2).style = 'besoin_fort'
        elif 'moyen' in niveau.lower():
            ws.cell(row=row, column=2).style = 'besoin_moyen'
        else:
            ws.cell(row=row, column=2).style = 'cellule'
        
        for col in [3, 4, 5, 6]:
            ws.cell(row=row, column=col).style = 'cellule'
            if col >= 4:
                ws.cell(row=row, column=col).font = Font(name='Wingdings', size=14)
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
        
        ws.row_dimensions[row].height = 20
        row += 1
    
    # Ligne totaux
    ws.cell(row=row, column=1, value="TOTAL")
    ws.cell(row=row, column=1).style = 'total'
    ws.cell(row=row, column=3, value=f"=SUM(C4:C{row-1})")
    ws.cell(row=row, column=3).style = 'total'
    
    # Ajustement colonnes
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    
    # Instructions
    ws.cell(row=row+2, column=1, value="Instructions : Cochez les cases selon l'état d'avancement. Double-cliquez sur une case pour la cocher/décocher.")
    ws.cell(row=row+2, column=1).font = Font(italic=True, size=9)


def creer_onglet_budget(ws, programme):
    """Crée l'onglet Budget prévisionnel avec formules"""
    # Titre
    ws.merge_cells('A1:D1')
    ws['A1'] = "BUDGET PRÉVISIONNEL"
    ws['A1'].style = 'titre'
    ws.row_dimensions[1].height = 30
    
    estimation = programme.get('estimation', {})
    
    # Section Coûts de formation
    ws['A3'] = "COÛTS DE FORMATION"
    ws['A3'].style = 'sous_titre'
    ws.merge_cells('A3:D3')
    
    data = [
        ["Poste", "Quantité", "Prix unitaire", "Total"],
        ["Heures de formation", estimation.get('total_heures', 0), estimation.get('cout_horaire', 45), "=B5*C5"],
        ["Nombre de séances", estimation.get('nb_seances', 1), "", ""],
        ["Durée moyenne par séance", "3,5h", "", ""],
    ]
    
    row = 4
    for ligne in data:
        for col, value in enumerate(ligne, 1):
            cell = ws.cell(row=row, column=col, value=value)
            if row == 4:
                cell.style = 'entete'
            else:
                cell.style = 'cellule'
                if col == 4 and row == 5:  # Formule total
                    cell.number_format = '#,##0.00 €'
        row += 1
    
    # Section Certification
    ws[f'A{row+1}'] = "CERTIFICATION"
    ws[f'A{row+1}'].style = 'sous_titre'
    ws.merge_cells(f'A{row+1}:D{row+1}')
    
    row += 2
    cert_data = [
        ["Poste", "", "Prix", ""],
        ["Certification TOSA", "", estimation.get('cout_certification', 0), ""],
    ]
    
    for ligne in cert_data:
        for col, value in enumerate(ligne, 1):
            cell = ws.cell(row=row, column=col, value=value)
            if ligne[0] == "Poste":
                cell.style = 'entete'
            else:
                cell.style = 'cellule'
                if col == 3:
                    cell.number_format = '#,##0.00 €'
        row += 1
    
    # Section Totaux
    ws[f'A{row+1}'] = "TOTAL ESTIMÉ"
    ws[f'A{row+1}'].style = 'sous_titre'
    ws.merge_cells(f'A{row+1}:D{row+1}')
    
    row += 2
    total_data = [
        ["Description", "Montant HT", "TVA (20%)", "Montant TTC"],
        ["Formation", "=D5", "=B{0}*0.2".format(row), "=B{0}+C{0}".format(row)],
        ["Certification", estimation.get('cout_certification', 0), "=B{0}*0.2".format(row+1), "=B{0}+C{0}".format(row+1)],
        ["TOTAL", "=SUM(B{0}:B{1})".format(row, row+1), "=SUM(C{0}:C{1})".format(row, row+1), "=SUM(D{0}:D{1})".format(row, row+1)],
    ]
    
    for i, ligne in enumerate(total_data):
        for col, value in enumerate(ligne, 1):
            cell = ws.cell(row=row+i, column=col, value=value)
            if i == 0:
                cell.style = 'entete'
            elif i == 3:  # Ligne total
                cell.style = 'total'
                cell.number_format = '#,##0.00 €'
            else:
                cell.style = 'cellule'
                cell.number_format = '#,##0.00 €'
    
    # Ajustement colonnes
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    
    # Protection formules
    for row_cells in ws[f'B5:D{row+3}']:
        for cell in row_cells:
            if cell.value and str(cell.value).startswith('='):
                cell.protection = Protection(locked=True)


def creer_onglet_ressources(ws, programme, base_dir):
    """Crée l'onglet Ressources avec liens vers fiches"""
    # Titre
    ws.merge_cells('A1:E1')
    ws['A1'] = "RESSOURCES PÉDAGOGIQUES"
    ws['A1'].style = 'titre'
    ws.row_dimensions[1].height = 30
    
    # En-têtes
    headers = ['Thème', 'Compétence', 'Fiche PDF', 'Lien', 'Priorité']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.style = 'entete'
    
    ws.row_dimensions[3].height = 25
    
    # Chemin vers les fiches
    docs_path = Path(base_dir) / "data" / "bd" / "docs" / "cours"
    
    # Données
    row = 4
    theme_index = 0
    for theme in programme.get('themes_a_former', []):
        theme_index += 1
        theme_nom = f"Thème {theme_index} : {theme['nom']}"
        niveau = theme.get('niveau_besoin', '')

        # Pour chaque compétence du thème
        for comp in theme.get('comp_par_domaine', []):
            domaine_nom = comp.get('nom_domaine', '')
            
            for competence in comp.get('competences', []):
                ws.cell(row=row, column=1, value=theme_nom)
                ws.cell(row=row, column=2, value=f"{domaine_nom} - {competence}")
                
                # Recherche de la fiche correspondante
                fiche_nom = trouver_fiche_correspondante(theme_nom, competence, docs_path)
                ws.cell(row=row, column=3, value=fiche_nom if fiche_nom else "Fiche générique")
                
                # Lien hypertexte si fiche existe
                if fiche_nom:
                    lien_cell = ws.cell(row=row, column=4, value="📄 Ouvrir")
                    lien_cell.style = 'lien'
                    lien_cell.hyperlink = str(docs_path / fiche_nom)
                else:
                    ws.cell(row=row, column=4, value="-")
                
                # Priorité
                if 'fort' in niveau.lower():
                    ws.cell(row=row, column=5, value="🔴 Haute")
                elif 'moyen' in niveau.lower():
                    ws.cell(row=row, column=5, value="� Moyenne")
                else:
                    ws.cell(row=row, column=5, value="🟢 Basse")
                
                # Styles
                for col in [1, 2, 3, 5]:
                    ws.cell(row=row, column=col).style = 'cellule'
                
                ws.row_dimensions[row].height = 20
                row += 1
    
    # Ajustement colonnes
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    
    # Note
    ws.cell(row=row+2, column=1, value="Note : Les liens fonctionnent si les fichiers PDF sont accessibles depuis cet emplacement.")
    ws.cell(row=row+2, column=1).font = Font(italic=True, size=9)


def trouver_fiche_correspondante(theme, competence, docs_path):
    """Trouve la fiche PDF correspondant au thème et compétence"""
    if not docs_path.exists():
        return None
    
    # Mapping simple basé sur les mots-clés
    mapping = {
        'prise en main': '1- Word Prise en main.pdf',
        'mise en forme simple': '1- Word Prise en main.pdf',
        'mise en forme': '2- Word PM compléments.pdf',
        'tabulation': '3- Word Tabulations.pdf',
        'tableau': '4- Word Tableaux.pdf',
        'présentation': '5- Word Présentations élaborées 365.pdf',
        'en-tête': '6- Word EnTetes et pieds de pages.pdf',
        'publipostage': '7- word publipostage initiation.pdf',
        'publipostage perfectionnement': '8- word publipostage perfectionnement.pdf',
        'longs documents': '9- word longs documents initiation.pdf',
        'longs documents perfectionnement': '10- word longs documents perfectionnement.pdf',
        'formulaires': '11- word Formulaires new.pdf',
        'champs': '12- word champs.pdf',
        'insertions automatiques': '13- word insertions automatiques.pdf',
        'modèles': '14- Word modèles.pdf',
        'excel vers word': '15- word Excel vers Word.pdf',
    }
    
    # Strip "Thème N : " prefix before matching
    import re
    theme_clean = re.sub(r'^thème\s*\d+\s*:\s*', '', theme_lower, flags=re.IGNORECASE)

    for key, value in mapping.items():
        if key in theme_clean:
            return value
    
    return None


def generer_excel_programme(id_candidat, base_dir, output_path=None):
    """
    Génère le fichier Excel complet pour un candidat
    
    Args:
        id_candidat: ID du candidat
        base_dir: Répertoire racine du projet
        output_path: Chemin de sortie (optionnel)
    
    Returns:
        Chemin du fichier Excel généré
    """
    # Chargement des données
    programme = charger_programme(id_candidat, base_dir)
    if not programme:
        raise ValueError(f"Programme non trouvé pour le candidat {id_candidat}")
    
    # Création du classeur
    wb = Workbook()
    creer_styles(wb)
    
    # Suppression de la feuille par défaut
    wb.remove(wb.active)
    
    candidat_info = {
        'nom': programme.get('nom_candidat', ''),
        'prenom': programme.get('prenom_candidat', '')
    }
    
    # Onglet 1: Planning
    ws_planning = wb.create_sheet("Planning")
    creer_onglet_planning(ws_planning, programme, candidat_info)
    
    # Onglet 2: Suivi
    ws_suivi = wb.create_sheet("Suivi progression")
    creer_onglet_suivi(ws_suivi, programme)
    
    # Onglet 3: Budget
    ws_budget = wb.create_sheet("Budget")
    creer_onglet_budget(ws_budget, programme)
    
    # Onglet 4: Ressources
    ws_ressources = wb.create_sheet("Ressources")
    creer_onglet_ressources(ws_ressources, programme, base_dir)
    
    # Détermination du chemin de sortie
    if not output_path:
        import os
        output_dir = Path("/tmp/posidata/export") if os.environ.get('VERCEL') == '1' else Path(base_dir) / "data" / "export"
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = output_dir / f"programme_{id_candidat}.xlsx"
    
    # Sauvegarde
    wb.save(output_path)
    
    return str(output_path)



# ==================== QUESTIONNAIRE TEMPLATE ====================

DOMAIN_NAMES = {
    'D1': 'Environnement / Méthodes',
    'D2': 'Calculs & Formules',
    'D3': 'Mise en forme',
    'D4': 'Gestion des données',
}

LEVEL_LABELS = {
    'QUEST_EXCEL_INIT':  ('Excel — Initiation',           'N1',    'Débutant complet'),
    'QUEST_EXCEL_INTER': ('Excel — Intermédiaire',         'N2',    'Pratique basique déjà acquise'),
    'QUEST_EXCEL_PERF':  ('Excel — Perfectionnement',      'N3',    'Niveau intermédiaire maîtrisé'),
    'QUEST_WORD':        ('Word — Rédaction professionnelle', 'N1-N3', 'Tous niveaux'),
}

C_NAVY      = '0F1F3D'
C_NAVY_MID  = '1A3361'
C_NAVY_LIGHT= '2D4A7A'
C_GOLD      = 'C9A84C'
C_GOLD_PALE = 'FDF8EE'
C_CREAM     = 'F7F5F0'
C_WHITE     = 'FFFFFF'
C_BORDER    = 'D0D5DD'
DOMAIN_COLORS  = {'D1': 'E8EDF6', 'D2': 'EAF3E0', 'D3': 'FEF3E2', 'D4': 'F0EAFA'}
DOMAIN_ACCENT  = {'D1': '2D4A7A', 'D2': '2E7D32', 'D3': '9C6C00', 'D4': '6A1B9A'}


def _brd(color=C_BORDER):
    s = Side(style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _fill2(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')


def _creer_onglet_garde(ws, questionnaire, id_questionnaire):
    label, niveau, profil = LEVEL_LABELS.get(id_questionnaire,
        (questionnaire.get('titre', id_questionnaire), '', ''))
    nb_q = len(questionnaire.get('questions', []))
    ws.sheet_view.showGridLines = False

    # Bandeau titre
    ws.merge_cells('A1:G3')
    c = ws['A1']
    c.value = f"Questionnaire de positionnement\n{label}"
    c.font = Font(name='Calibri', size=18, bold=True, color=C_WHITE)
    c.fill = _fill2(C_NAVY)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for r in [1, 2, 3]:
        ws.row_dimensions[r].height = 26

    # Sous-bandeau doré
    ws.merge_cells('A4:G4')
    c = ws['A4']
    c.value = f"Niveau : {niveau}  ·  Profil : {profil}  ·  {nb_q} compétences à évaluer"
    c.font = Font(name='Calibri', size=10, bold=True, color=C_NAVY)
    c.fill = _fill2(C_GOLD)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[4].height = 20

    def section_hdr(row, txt):
        ws.merge_cells(f'A{row}:G{row}')
        c = ws[f'A{row}']
        c.value = txt
        c.font = Font(name='Calibri', size=11, bold=True, color=C_WHITE)
        c.fill = _fill2(C_NAVY_MID)
        c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.row_dimensions[row].height = 22

    def info_row(row, lbl, placeholder=''):
        ws.merge_cells(f'B{row}:G{row}')
        c = ws[f'A{row}']
        c.value = lbl
        c.font = Font(name='Calibri', size=10, bold=True)
        c.fill = _fill2(C_CREAM)
        c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        c2 = ws[f'B{row}']
        c2.value = placeholder
        c2.font = Font(name='Calibri', size=10, italic=True, color='888888')
        c2.fill = _fill2(C_WHITE)
        c2.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = _brd()
        ws.row_dimensions[row].height = 20

    section_hdr(6, '📋  Informations du candidat')
    info_row(7, 'Nom :', 'Saisir le nom de famille')
    info_row(8, 'Prénom :', 'Saisir le prénom')
    info_row(9, 'Date :', datetime.now().strftime('%d/%m/%Y'))
    info_row(10, 'Formateur(trice) :', '')
    info_row(11, 'Organisme :', 'eS@deq')

    section_hdr(13, "📖  Mode d'emploi")

    instructions = [
        ('1', "Rendez-vous sur l'onglet « 📝 Questionnaire » (onglet suivant)."),
        ('2', "Pour chaque compétence, évaluez votre niveau dans la colonne « Maîtrise actuelle » :\n"
              "    • Aucune  → Je ne connais pas du tout\n"
              "    • Moyenne → Je connais les bases mais manque de pratique\n"
              "    • Acquise → Je maîtrise sans difficulté"),
        ('3', "Dans « Souhait de formation » indiquez :\n"
              "    • Oui → Je souhaite être formé(e) sur ce point\n"
              "    • Non → Ce point ne m\'est pas utile / je suis déjà autonome"),
        ('4', "Astuce : cliquez sur ▼ dans la cellule pour choisir dans la liste — ne saisissez pas manuellement."),
        ('5', "Une fois terminé, enregistrez le fichier et remettez-le à votre formateur(trice).\n"
              "Il sera importé dans la plateforme pour générer votre programme personnalisé."),
    ]

    row = 14
    for num, txt in instructions:
        nb_lines = txt.count('\n') + 1
        end_row = row + nb_lines - 1
        ws.merge_cells(f'B{row}:G{end_row}')
        ws[f'A{row}'].value = num
        ws[f'A{row}'].font = Font(name='Calibri', size=12, bold=True, color=C_WHITE)
        ws[f'A{row}'].fill = _fill2(C_GOLD)
        ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='top')
        ws[f'B{row}'].value = txt
        ws[f'B{row}'].font = Font(name='Calibri', size=10)
        ws[f'B{row}'].fill = _fill2(C_WHITE)
        ws[f'B{row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        for r in range(row, end_row + 1):
            ws.row_dimensions[r].height = 16
            for col in range(1, 8):
                ws.cell(row=r, column=col).border = _brd()
        row = end_row + 1

    ws.merge_cells(f'A{row+1}:G{row+1}')
    c = ws[f'A{row+1}']
    c.value = '🔒  Vos réponses sont strictement confidentielles et utilisées uniquement pour personnaliser votre parcours.'
    c.font = Font(name='Calibri', size=9, italic=True, color='555555')
    c.fill = _fill2(C_GOLD_PALE)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[row + 1].height = 24

    ws.column_dimensions['A'].width = 22
    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col_letter].width = 15


def _creer_onglet_questionnaire(ws, questionnaire, id_questionnaire):
    import re
    ws.sheet_view.showGridLines = False
    label, niveau, _ = LEVEL_LABELS.get(id_questionnaire,
        (questionnaire.get('titre', id_questionnaire), '', ''))

    # En-tête
    ws.merge_cells('A1:F1')
    c = ws['A1']
    c.value = f"Questionnaire — {label}  |  Niveau {niveau}"
    c.font = Font(name='Calibri', size=13, bold=True, color=C_WHITE)
    c.fill = _fill2(C_NAVY)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:F2')
    c = ws['A2']
    c.value = "Complétez les colonnes « Maîtrise actuelle » et « Souhait de formation » à l'aide des listes déroulantes ▼"
    c.font = Font(name='Calibri', size=9, italic=True, color=C_NAVY)
    c.fill = _fill2(C_GOLD_PALE)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 16

    HEADERS = ['#', 'Compétence à évaluer', 'id_question', 'id_questionnaire',
               'Maîtrise actuelle', 'Souhait de formation']
    WIDTHS  = [5, 60, 16, 20, 20, 22]
    for col, (h, w) in enumerate(zip(HEADERS, WIDTHS), 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = Font(name='Calibri', size=10, bold=True, color=C_WHITE)
        cell.fill = _fill2(C_NAVY_LIGHT)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = _brd(C_NAVY)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[3].height = 28
    ws.freeze_panes = 'A4'

    # Regrouper par domaine
    domaines_dict = {}
    for q in questionnaire.get('questions', []):
        m = re.match(r'\[?(D\d+)-', q['id_question'])
        if m:
            dom = m.group(1)
        else:
            dom = q.get('id_theme', 'GEN')[:2]  # T1, T2… pour Word
        if dom not in domaines_dict:
            domaines_dict[dom] = []
        domaines_dict[dom].append(q)

    # Validations déroulantes
    dv_maitrise = DataValidation(type='list', formula1='"Aucune,Moyenne,Acquise"',
                                  allow_blank=True, showDropDown=False)
    dv_besoin   = DataValidation(type='list', formula1='"Oui,Non"',
                                  allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv_maitrise)
    ws.add_data_validation(dv_besoin)

    row = 4
    q_num = 0
    for dom_id in sorted(domaines_dict.keys()):
        questions = domaines_dict[dom_id]
        dom_name = DOMAIN_NAMES.get(dom_id, f'Thème {dom_id}')
        dom_color = DOMAIN_COLORS.get(dom_id, 'E0E0E0')
        dom_accent = DOMAIN_ACCENT.get(dom_id, C_NAVY_MID)

        # Bandeau domaine
        ws.merge_cells(f'A{row}:F{row}')
        c = ws[f'A{row}']
        c.value = f"  ● {dom_name}    ({len(questions)} compétence{'s' if len(questions) > 1 else ''})"
        c.font = Font(name='Calibri', size=11, bold=True, color=C_WHITE)
        c.fill = _fill2(dom_accent)
        c.alignment = Alignment(horizontal='left', vertical='center')
        c.border = Border(top=Side(style='medium', color=C_NAVY),
                          bottom=Side(style='medium', color=C_NAVY))
        ws.row_dimensions[row].height = 22
        row += 1

        for i, q in enumerate(questions):
            q_num += 1
            bg = C_WHITE if i % 2 == 0 else C_CREAM

            vals = [q_num, q.get('texte', ''), q['id_question'], id_questionnaire, 'Aucune', 'Oui']
            for col, val in enumerate(vals, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.font = Font(name='Calibri', size=10)
                c.fill = _fill2(bg)
                c.border = _brd()
                c.alignment = Alignment(horizontal='center' if col != 2 else 'left',
                                        vertical='center', wrap_text=(col == 2))

            # Styles spéciaux
            ws.cell(row=row, column=1).font = Font(name='Calibri', size=9, color='888888')
            ws.cell(row=row, column=3).font = Font(name='Courier New', size=9, color='444444')
            ws.cell(row=row, column=4).font = Font(name='Courier New', size=9, color='AAAAAA')
            ws.cell(row=row, column=5).font = Font(name='Calibri', size=10, bold=True, color=C_NAVY)
            ws.cell(row=row, column=6).font = Font(name='Calibri', size=10, bold=True, color='8B6914')

            dv_maitrise.add(ws.cell(row=row, column=5))
            dv_besoin.add(ws.cell(row=row, column=6))

            ws.row_dimensions[row].height = 22
            row += 1

    # Pied de page
    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    c = ws[f'A{row}']
    c.value = (f'Questionnaire généré le {datetime.now().strftime("%d/%m/%Y à %Hh%M")} '
               f'— eS@deq Plateforme de Positionnement')
    c.font = Font(name='Calibri', size=8, italic=True, color='AAAAAA')
    c.alignment = Alignment(horizontal='right')
    ws.row_dimensions[row].height = 14

    ws.print_title_rows = '1:3'
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1


def generer_excel_questionnaire(id_questionnaire, base_dir, output_path=None):
    """Génère un fichier Excel stylé questionnaire type pour candidats."""
    q_path = Path(base_dir) / 'data' / 'bd' / '00_questions.json'
    with open(q_path, 'r', encoding='utf-8') as f:
        questions_data = json.load(f)

    questionnaire = next(
        (q for q in questions_data.get('questionnaires', [])
         if q['id_questionnaire'] == id_questionnaire),
        None
    )
    if not questionnaire:
        raise ValueError(f"Questionnaire {id_questionnaire} non trouvé")

    wb = Workbook()
    wb.remove(wb.active)

    ws_garde = wb.create_sheet('📋 Accueil')
    _creer_onglet_garde(ws_garde, questionnaire, id_questionnaire)

    ws_quest = wb.create_sheet('📝 Questionnaire')
    _creer_onglet_questionnaire(ws_quest, questionnaire, id_questionnaire)

    wb.active = ws_quest

    if not output_path:
        import os
        output_dir = Path("/tmp/posidata/export") if os.environ.get('VERCEL') == '1' else Path(base_dir) / 'data' / 'export'
        output_dir.mkdir(parents=True, exist_ok=True)
        safe_id = id_questionnaire.lower().replace('quest_', '')
        output_path = output_dir / f'questionnaire_type_{safe_id}.xlsx'

    wb.save(output_path)
    return str(output_path)


if __name__ == "__main__":
    # Test
    import sys
    if len(sys.argv) > 1:
        candidat_id = sys.argv[1]
    else:
        candidat_id = "CAND_S_o_20260221_002154"
    
    base = r"C:\laragon\www\test_ia\posi-2026-v2"
    chemin = generer_excel_programme(candidat_id, base)
    print(f"Excel généré : {chemin}")
