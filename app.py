"""
Application Flask - eS@deq Plateforme de Positionnement
Architecture v2 — Février 2026
"""
import json
import csv
import os
import shutil
import io
from datetime import datetime
from pathlib import Path
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file
import config

app = Flask(__name__)
app.config['SECRET_KEY'] = config.SECRET_KEY
app.config['DEBUG'] = config.DEBUG
app.config['APP_VERSION'] = config.APP_VERSION
app.config['LAST_UPDATE'] = config.LAST_UPDATE

# Import des modules locaux
import sys
sys.path.insert(0, str(config.BASE_DIR))
from data.analyse_besoins import (
    creer_programme_personnalise, generer_programme,
    analyser_reponse, charger_reponses_candidat, grouper_reponses_par_domaine,
    grouper_reponses_par_bloc,
    REGLE_BESOIN_FORT, REGLE_BESOIN_MOYEN, REGLE_A_REVOIR, REGLE_IGNORER
)


def load_json_file(filepath):
    """Charge un fichier JSON"""
    with open(filepath, 'r', encoding='utf-8') as f:
        return json.load(f)


def save_json_file(filepath, data):
    """Sauvegarde un fichier JSON"""
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def allowed_file(filename):
    """Vérifie si le fichier a une extension autorisée (csv, txt, xlsx)"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in {'csv', 'txt', 'xlsx'}


def build_question_text_lookup(questions_data):
    """
    Construit un dictionnaire {id_question: texte} depuis 00_questions.json.
    Pour enrichir les réponses avec le texte de la question (traçabilité).
    """
    lookup = {}
    for quest in questions_data.get('questionnaires', []):
        for q in quest.get('questions', []):
            lookup[q['id_question']] = q.get('texte', '')
    return lookup


def parse_csv_candidat(filepath):
    """
    Parse un fichier CSV de positionnement candidat.
    Format attendu: id_questionnaire;id_question;acquisition;besoin
    Retourne (id_questionnaire, reponses[]) ou (None, [])
    """
    reponses = []
    id_questionnaire = None

    for delimiter in [';', ',']:
        try:
            with open(filepath, 'r', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f, delimiter=delimiter)
                fieldnames = reader.fieldnames or []
                if 'id_question' not in fieldnames:
                    continue
                for row in reader:
                    id_q = row.get('id_question', '').strip()
                    if not id_q:
                        continue
                    if not id_questionnaire and 'id_questionnaire' in row:
                        id_questionnaire = row['id_questionnaire'].strip()
                    reponses.append({
                        'id_question': id_q,
                        'acquisition': row.get('acquisition', '').strip(),
                        'besoin': row.get('besoin', '').strip()
                    })
                if reponses:
                    break
        except Exception:
            continue

    return id_questionnaire, reponses


def parse_excel_questionnaire(filepath):
    """
    Parse un fichier Excel généré par generer_excel_questionnaire().
    Lit l'onglet '📝 Questionnaire', colonnes :
      C (3) = id_question, D (4) = id_questionnaire,
      E (5) = Maîtrise actuelle, F (6) = Souhait de formation
    Ignore les lignes de bandeau (cellules fusionnées / sans id_question valide).
    Retourne (id_questionnaire, reponses[]) ou (None, [])
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError("openpyxl est requis pour importer les fichiers Excel")

    wb = load_workbook(filepath, data_only=True)

    # Trouver l'onglet questionnaire ("📝 Questionnaire" ou premier onglet avec données)
    ws = None
    for sheet_name in wb.sheetnames:
        if 'questionnaire' in sheet_name.lower() or 'Quest' in sheet_name:
            ws = wb[sheet_name]
            break
    if ws is None:
        ws = wb.active

    reponses = []
    id_questionnaire = None

    # Parcourir à partir de la ligne 4 (après les 3 lignes d'en-tête)
    for row in ws.iter_rows(min_row=4, values_only=True):
        # Colonnes : A=0, B=1, C=2(id_q), D=3(id_quest), E=4(maitrise), F=5(besoin)
        if len(row) < 6:
            continue
        id_q = str(row[2]).strip() if row[2] is not None else ''
        id_quest = str(row[3]).strip() if row[3] is not None else ''
        maitrise = str(row[4]).strip() if row[4] is not None else ''
        besoin   = str(row[5]).strip() if row[5] is not None else ''

        # Ignorer les lignes vides ou bandeaux de domaine
        if not id_q or id_q in ('None', 'id_question') or id_q.startswith('●'):
            continue
        # Normaliser Maîtrise (Aucune → Aucun pour compatibilité analyseur)
        maitrise_norm = {'Aucune': 'Aucun', 'Moyenne': 'Moyen', 'Acquise': 'Acquis'}.get(maitrise, maitrise)

        if not id_questionnaire and id_quest:
            id_questionnaire = id_quest

        reponses.append({
            'id_question': id_q,
            'acquisition': maitrise_norm,
            'besoin': besoin
        })

    return id_questionnaire, reponses


# ==================== FILTERS ====================

@app.template_filter('format_fr')
def format_fr(value):
    """Formatte un nombre au format français (virgule décimale, espace milliers)"""
    try:
        if value is None: return "0,00"
        # Utiliser un espace insécable pour les milliers
        return f"{float(value):,.2f}".replace(",", " ").replace(".", ",").replace(" ", "\u00A0")
    except (ValueError, TypeError):
        return value

@app.template_filter('format_integer_fr')
def format_integer_fr(value):
    """Formatte un entier au format français"""
    try:
        if value is None: return "0"
        return f"{int(value):,}".replace(",", "\u00A0")
    except (ValueError, TypeError):
        return value

@app.template_filter('format_date_fr')
def format_date_fr(value):
    """Formatte une date string en jj/mm/aaaa"""
    if not value: return ""
    try:
        for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%Y%m%d%H%M%S'):
            try:
                dt = datetime.strptime(str(value), fmt)
                return dt.strftime('%d/%m/%Y')
            except ValueError:
                continue
        return value
    except Exception:
        return value

# ==================== ROUTES ====================

@app.route('/')
def index():
    """Page d'accueil - redirection vers dashboard"""
    return redirect(url_for('dashboard'))


@app.route('/favicon.ico')
def favicon():
    """Favicon optionnel pour éviter les erreurs 404 navigateur."""
    favicon_path = Path(app.root_path) / 'static' / 'favicon.ico'
    if favicon_path.exists():
        return send_file(favicon_path)
    return ('', 204)


@app.route('/dashboard')
def dashboard():
    """Dashboard avec statistiques globales"""
    try:
        candidats_data = load_json_file(config.CANDIDATS_FILE)
        candidats = candidats_data.get('candidats', [])

        # Statistiques
        total_candidats = len(candidats)
        stats_besoins = {'besoin_fort': 0, 'besoin_moyen': 0, 'a_revoir': 0, 'aucun': 0}

        # Compter les programmes générés
        total_programmes = 0

        # Load questions globally to avoid loop lookup
        questions_data = load_json_file(config.QUESTIONS_FILE)
        
        # Analyser les réponses depuis les dossiers candidats
        for candidat in candidats:
            id_cand = candidat['id_candidat']
            questionnaire = charger_reponses_candidat(id_cand)
            if questionnaire:
                self_level = questionnaire.get('self_level', candidat.get('self_level', ''))
                if self_level == 'novice':
                    quest_id = candidat.get('id_questionnaire') or 'QUEST_EXCEL_INIT'
                    quest_def = next((q for q in questions_data.get('questionnaires', []) if q['id_questionnaire'] == quest_id), None)
                    nb_max = len(quest_def.get('questions', [])) if quest_def else 1
                    stats_besoins['besoin_fort'] += nb_max
                else:
                    for reponse in questionnaire.get('reponses', []):
                        type_besoin = analyser_reponse(reponse.get('acquisition'), reponse.get('besoin'))
                        if type_besoin == REGLE_IGNORER:
                            stats_besoins['aucun'] += 1
                        elif type_besoin == REGLE_BESOIN_FORT:
                            stats_besoins['besoin_fort'] += 1
                        elif type_besoin == REGLE_BESOIN_MOYEN:
                            stats_besoins['besoin_moyen'] += 1
                        elif type_besoin == REGLE_A_REVOIR:
                            stats_besoins['a_revoir'] += 1

        # Vérifier les programmes existants
        programmes_dir = config.CANDIDATS_DIR
        if programmes_dir.exists():
            for cand_dir in programmes_dir.iterdir():
                if cand_dir.is_dir() and (cand_dir / 'programme_perso.json').exists():
                    total_programmes += 1

        return render_template('dashboard.html',
                             total_candidats=total_candidats,
                             stats_besoins=stats_besoins,
                             total_programmes=total_programmes,
                             candidats=candidats[:5])  # 5 derniers pour apercu
    except Exception as e:
        flash(f"Erreur lors du chargement du dashboard: {str(e)}", 'error')
        return render_template('dashboard.html', total_candidats=0, stats_besoins={}, total_programmes=0, candidats=[])


@app.route('/candidats')
def liste_candidats():
    """Liste des candidats avec pagination et filtres"""
    try:
        candidats_data = load_json_file(config.CANDIDATS_FILE)
        candidats = candidats_data.get('candidats', [])

        # Trier par date (plus récents d'abord)
        candidats = sorted(candidats, key=lambda x: x.get('date_reponse', ''), reverse=True)

        # Vérifier si chaque candidat a un programme, et analyser ses besoins
        for cand in candidats:
            prog_path = config.CANDIDATS_DIR / cand['id_candidat'] / 'programme_perso.json'
            cand['programme_genere'] = prog_path.exists()
            cand['has_fort'] = False
            cand['has_moyen'] = False
            if prog_path.exists():
                try:
                    with open(prog_path, 'r', encoding='utf-8') as f:
                        p = json.load(f)
                        needs = [t.get('niveau_besoin') for t in p.get('themes_a_former', [])]
                        if 'fort' in needs: cand['has_fort'] = True
                        if 'moyen' in needs: cand['has_moyen'] = True
                except:
                    pass

        # Filtrage par type
        filtrer = request.args.get('filter', '')
        if filtrer == 'programmes':
            candidats = [c for c in candidats if c['programme_genere']]
        elif filtrer == 'sans_programme':
            candidats = [c for c in candidats if not c['programme_genere']]
        elif filtrer == 'forts':
            candidats = [c for c in candidats if c['has_fort']]
        elif filtrer == 'moyens':
            candidats = [c for c in candidats if c['has_moyen']]

        # Filtrage Temporel (esadeq_timeTravel)
        date_filter = request.args.get('date_filter', 'all')
        if date_filter != 'all':
            from datetime import timedelta
            now = datetime.now()
            filtered_candidats = []
            
            for c in candidats:
                date_str = c.get('date_reponse', '')
                if not date_str:
                    continue
                try:
                    c_date = datetime.strptime(date_str, '%Y-%m-%d %H:%M:%S')
                    # Appliquer le filtre
                    if date_filter == 'today':
                        if c_date.date() == now.date():
                            filtered_candidats.append(c)
                    elif date_filter == 'week':
                        start_of_week = now.date() - timedelta(days=now.weekday())
                        if c_date.date() >= start_of_week:
                            filtered_candidats.append(c)
                    elif date_filter == 'month':
                        if c_date.year == now.year and c_date.month == now.month:
                            filtered_candidats.append(c)
                    elif date_filter == 'year':
                        if c_date.year == now.year:
                            filtered_candidats.append(c)
                except ValueError:
                    pass
            candidats = filtered_candidats

        # Recherche textuelle
        q = request.args.get('q', '').strip().lower()
        if q:
            candidats = [c for c in candidats if q in c.get('nom', '').lower() or q in c.get('prenom', '').lower() or q in c.get('id_candidat', '').lower()]

        # Pagination simple
        page = request.args.get('page', 1, type=int)
        per_page = config.ITEMS_PER_PAGE
        total = len(candidats)
        start = (page - 1) * per_page
        end = start + per_page
        candidats_page = candidats[start:end]

        return render_template('candidat/liste.html',
                             candidats=candidats_page,
                             page=page,
                             total=total,
                             per_page=per_page,
                             current_filter=filtrer,
                             current_date_filter=date_filter,
                             current_q=q)
    except Exception as e:
        flash(f"Erreur: {str(e)}", 'error')
        return render_template('candidat/liste.html', candidats=[], page=1, total=0, current_filter='', current_date_filter='all', current_q='')


@app.route('/candidats/import', methods=['GET', 'POST'])
def import_candidat():
    """Import d'un nouveau candidat via CSV (format v2)"""
    programmes_data = load_json_file(config.PROGRAMMES_FILE)
    programmes = programmes_data.get('programmes', [])

    # Charger les questionnaires pour le mapping questionnaire→programme
    questions_data = load_json_file(config.QUESTIONS_FILE)
    questionnaires = questions_data.get('questionnaires', [])
    quest_to_prog = {q['id_questionnaire']: q.get('id_programme') for q in questionnaires}

    if request.method == 'POST':
        if 'file' not in request.files:
            flash('Aucun fichier sélectionné', 'error')
            return redirect(request.url)

        file = request.files['file']
        if file.filename == '':
            flash('Aucun fichier sélectionné', 'error')
            return redirect(request.url)

        if file and allowed_file(file.filename):
            try:
                ext = file.filename.rsplit('.', 1)[1].lower()
                timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
                id_candidat = f"CAND_{timestamp}"

                nom = request.form.get('nom', '').strip()
                prenom = request.form.get('prenom', '').strip()

                if not nom:
                    flash('Le nom est obligatoire', 'error')
                    return redirect(request.url)

                # Sauvegarder le fichier uploadé temporairement
                tmp_filename = f"{id_candidat}.{ext}"
                filepath = config.UPLOAD_DIR / tmp_filename
                file.save(filepath)

                # Parser selon le format
                if ext == 'xlsx':
                    id_questionnaire, reponses = parse_excel_questionnaire(filepath)
                    msg_format = 'Excel'
                else:
                    id_questionnaire, reponses = parse_csv_candidat(filepath)
                    msg_format = 'CSV'

                if not reponses:
                    flash(f'Aucune réponse valide trouvée dans le fichier {msg_format}', 'error')
                    os.remove(filepath)
                    return redirect(request.url)

                # Si nom/prénom non saisis, tenter de les lire depuis l'onglet Accueil (xlsx)
                if ext == 'xlsx' and (not nom or not prenom):
                    try:
                        from openpyxl import load_workbook as _lw
                        _wb = _lw(filepath, data_only=True)
                        for _sn in _wb.sheetnames:
                            if 'accueil' in _sn.lower():
                                _ws = _wb[_sn]
                                nom_cell    = _ws['B7'].value
                                prenom_cell = _ws['B8'].value
                                if nom_cell and str(nom_cell).strip() not in ('Saisir le nom de famille', ''):
                                    nom = str(nom_cell).strip().upper()
                                if prenom_cell and str(prenom_cell).strip() not in ('Saisir le prénom', ''):
                                    prenom = str(prenom_cell).strip()
                                break
                    except Exception:
                        pass

                # Enrichir avec le texte des questions
                q_texts = build_question_text_lookup(questions_data)
                for rep in reponses:
                    rep['texte'] = q_texts.get(rep['id_question'], '')

                # Déduire le programme
                id_programme = quest_to_prog.get(id_questionnaire, 'PROG_WORD')

                # Créer le dossier candidat
                candidat_dir = config.CANDIDATS_DIR / id_candidat
                candidat_dir.mkdir(parents=True, exist_ok=True)

                # Sauvegarder les réponses (JSON)
                questionnaire_data = {
                    'id_candidat': id_candidat,
                    'id_questionnaire': id_questionnaire or 'QUEST_WORD',
                    'date_reponse': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'reponses': reponses
                }
                save_json_file(candidat_dir / 'questionnaire.json', questionnaire_data)

                # Archiver le fichier source
                dest_ext = 'xlsx' if ext == 'xlsx' else 'csv'
                os.rename(filepath, candidat_dir / f'questionnaire.{dest_ext}')

                # Ajouter à l'index candidats
                candidats_data = load_json_file(config.CANDIDATS_FILE)
                candidats_data['candidats'].append({
                    'id_candidat': id_candidat,
                    'nom': nom,
                    'prenom': prenom,
                    'id_questionnaire': id_questionnaire or 'QUEST_WORD',
                    'id_programme': id_programme,
                    'date_reponse': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                })
                save_json_file(config.CANDIDATS_FILE, candidats_data)

                # Lancer l'analyse automatiquement après l'import
                generer_programme(id_candidat, candidat_dir / 'programme_perso.json')

                prog_nom = next((p['intitule'] for p in programmes if p['id_programme'] == id_programme), id_programme)
                flash(f'Candidat {prenom} {nom} importé avec succès ({msg_format}) pour « {prog_nom} » !', 'success')
                return redirect(url_for('detail_candidat', id_candidat=id_candidat))

            except Exception as e:
                flash(f"Erreur lors de l'import : {str(e)}", 'error')
                return redirect(request.url)

    return render_template('candidat/import.html', programmes=programmes)


@app.route('/candidats/<id_candidat>')
def detail_candidat(id_candidat):
    """Détail d'un candidat avec son programme"""
    try:
        candidats_data = load_json_file(config.CANDIDATS_FILE)
        candidats = candidats_data.get('candidats', [])

        candidat = next((c for c in candidats if c['id_candidat'] == id_candidat), None)

        if not candidat:
            flash('Candidat non trouvé', 'error')
            return redirect(url_for('liste_candidats'))

        # Charger le programme s'il existe
        programme = None
        prog_path = config.CANDIDATS_DIR / id_candidat / 'programme_perso.json'
        if prog_path.exists():
            with open(prog_path, 'r', encoding='utf-8') as f:
                programme = json.load(f)
            # Régénérer si comp_par_domaine vide (migration)
            themes = programme.get('themes_a_former', [])
            if themes and not themes[0].get('comp_par_domaine'):
                programme = generer_programme(id_candidat, prog_path)

        # Charger les réponses depuis le dossier candidat
        questionnaire = charger_reponses_candidat(id_candidat)
        reponses = questionnaire.get('reponses', []) if questionnaire else []

        # Calculer les statistiques
        stats = {'fort': 0, 'moyen': 0, 'a_revoir': 0, 'total': len(reponses)}
        
        self_level = candidat.get('self_level', '')
        if questionnaire:
            self_level = questionnaire.get('self_level', self_level)
            
        if self_level == 'novice':
            questions_data = load_json_file(config.QUESTIONS_FILE)
            quest_id = candidat.get('id_questionnaire') or 'QUEST_EXCEL_INIT'
            quest_def = next((q for q in questions_data.get('questionnaires', []) if q['id_questionnaire'] == quest_id), None)
            nb_max = len(quest_def.get('questions', [])) if quest_def else 1
            stats['fort'] = nb_max
            stats['total'] = nb_max
        else:
            for reponse in reponses:
                type_besoin = analyser_reponse(reponse.get('acquisition'), reponse.get('besoin'))
                if type_besoin == REGLE_BESOIN_FORT:
                    stats['fort'] += 1
                elif type_besoin == REGLE_BESOIN_MOYEN:
                    stats['moyen'] += 1
                elif type_besoin == REGLE_A_REVOIR:
                    stats['a_revoir'] += 1

        # Grouper les réponses par domaine et par bloc pour l'affichage
        reponses_par_domaine = None
        reponses_par_bloc = None
        
        if programme:
            reponses_par_domaine = programme.get('reponses_par_domaine')
            reponses_par_bloc = programme.get('reponses_par_bloc')
            
        if not reponses_par_domaine:
            reponses_par_domaine = grouper_reponses_par_domaine(id_candidat, reponses)
        if not reponses_par_bloc:
            reponses_par_bloc = grouper_reponses_par_bloc(id_candidat, reponses)

        return render_template('candidat/detail.html',
                             candidat=candidat,
                             programme=programme,
                             reponses=reponses,
                             reponses_par_domaine=reponses_par_domaine,
                             reponses_par_bloc=reponses_par_bloc,
                             stats=stats)
    except Exception as e:
        flash(f"Erreur: {str(e)}", 'error')
        return redirect(url_for('liste_candidats'))


@app.route('/candidats/<id_candidat>/analyser')
def analyser_candidat(id_candidat):
    """Lance l'analyse et génère le programme pour un candidat"""
    try:
        # Générer le programme
        candidat_dir = config.CANDIDATS_DIR / id_candidat
        candidat_dir.mkdir(parents=True, exist_ok=True)
        fichier_sortie = candidat_dir / 'programme_perso.json'

        programme = generer_programme(id_candidat, fichier_sortie)

        if programme:
            flash('Programme généré avec succès!', 'success')
        else:
            flash('Erreur lors de la génération du programme. Vérifiez que le questionnaire existe.', 'error')

        return redirect(url_for('detail_candidat', id_candidat=id_candidat))
    except Exception as e:
        flash(f"Erreur lors de l'analyse: {str(e)}", 'error')
        return redirect(url_for('detail_candidat', id_candidat=id_candidat))


@app.route('/candidats/<id_candidat>/pdf')
def telecharger_pdf(id_candidat):
    """Génère et télécharge le PDF du programme"""
    try:
        # Import dynamique pour éviter erreur au démarrage
        from src.pdf_generator import generer_pdf_programme, is_pdf_available

        if not is_pdf_available():
            flash('La génération PDF n\'est pas disponible. Installez GTK3 pour activer cette fonctionnalité.', 'error')
            return redirect(url_for('detail_candidat', id_candidat=id_candidat))

        # Charger le programme
        prog_path = config.CANDIDATS_DIR / id_candidat / 'programme_perso.json'
        if not prog_path.exists():
            flash('Aucun programme généré. Lancez d\'abord l\'analyse.', 'error')
            return redirect(url_for('detail_candidat', id_candidat=id_candidat))

        with open(prog_path, 'r', encoding='utf-8') as f:
            programme = json.load(f)

        # Charger les infos candidat
        candidats_data = load_json_file(config.CANDIDATS_FILE)
        candidat = next((c for c in candidats_data['candidats'] if c['id_candidat'] == id_candidat), None)

        # Générer le PDF
        pdf_path = config.EXPORT_DIR / f"programme_{id_candidat}.pdf"
        pdf_bytes = generer_pdf_programme(programme, candidat, pdf_path)

        return send_file(
            io.BytesIO(pdf_bytes),
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f"programme_{id_candidat}.pdf"
        )
    except Exception as e:
        flash(f"Erreur lors de la génération PDF: {str(e)}", 'error')
        return redirect(url_for('detail_candidat', id_candidat=id_candidat))


@app.route('/candidats/<id_candidat>/excel')
def telecharger_excel(id_candidat):
    """Génère et télécharge le fichier Excel du programme"""
    try:
        # Import du générateur Excel
        from src.excel_generator import generer_excel_programme

        # Charger les infos candidat
        candidats_data = load_json_file(config.CANDIDATS_FILE)
        candidat = next((c for c in candidats_data['candidats'] if c['id_candidat'] == id_candidat), None)
        
        if not candidat:
            flash('Candidat non trouvé', 'error')
            return redirect(url_for('liste_candidats'))

        # Vérifier que le programme existe
        prog_path = config.CANDIDATS_DIR / id_candidat / 'programme_perso.json'
        if not prog_path.exists():
            flash('Aucun programme généré. Lancez d\'abord l\'analyse.', 'error')
            return redirect(url_for('detail_candidat', id_candidat=id_candidat))

        # Générer le fichier Excel
        base_dir = str(config.BASE_DIR)
        excel_path = generer_excel_programme(id_candidat, base_dir)
        
        # Envoyer le fichier
        return send_file(
            excel_path,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"programme_{id_candidat}.xlsx"
        )
        
    except ImportError as e:
        flash(f'Module Excel requis: pip install openpyxl. Erreur: {str(e)}', 'error')
        return redirect(url_for('detail_candidat', id_candidat=id_candidat))
    except Exception as e:
        flash(f"Erreur lors de la génération Excel: {str(e)}", 'error')
        return redirect(url_for('detail_candidat', id_candidat=id_candidat))


@app.route('/candidats/<id_candidat>/supprimer', methods=['POST'])
def supprimer_candidat(id_candidat):
    """Supprime un candidat actif et archive ses donnees."""
    try:
        candidats_data = load_json_file(config.CANDIDATS_FILE)
        candidats = candidats_data.get('candidats', [])
        candidat = next((c for c in candidats if c.get('id_candidat') == id_candidat), None)

        if not candidat:
            flash('Candidat non trouve', 'error')
            return redirect(url_for('liste_candidats'))

        archive_timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        archive_dir = config.ARCHIVES_CANDIDATS_DIR / f"{id_candidat}_{archive_timestamp}"
        archive_dir.mkdir(parents=True, exist_ok=True)

        candidat_dir = config.CANDIDATS_DIR / id_candidat
        archived_candidat_dir = archive_dir / 'candidat_data'
        if candidat_dir.exists():
            shutil.move(str(candidat_dir), str(archived_candidat_dir))

        pdf_file = config.EXPORT_DIR / f"programme_{id_candidat}.pdf"
        archived_pdf_file = archive_dir / pdf_file.name
        if pdf_file.exists():
            shutil.move(str(pdf_file), str(archived_pdf_file))

        archive_metadata = {
            'archive_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'id_candidat': id_candidat,
            'action': 'suppression_avec_archivage',
            'candidat': candidat,
            'archive_dir': str(archive_dir),
            'contenu': {
                'dossier_candidat_archive': archived_candidat_dir.exists(),
                'pdf_archive': archived_pdf_file.exists()
            }
        }
        save_json_file(archive_dir / 'archive_metadata.json', archive_metadata)

        candidats_data['candidats'] = [c for c in candidats if c.get('id_candidat') != id_candidat]
        flash(f"Candidat {id_candidat} supprime et archive avec succes.", 'success')
        return redirect(url_for('liste_candidats'))
    except Exception as e:
        app.logger.error(f"Erreur lors de la suppression: {str(e)}")
        flash(f"Erreur lors de la suppression: {str(e)}", 'error')
        return redirect(url_for('liste_candidats'))

@app.route('/candidats/batch', methods=['POST'])
def batch_action():
    action = request.form.get('action')
    ids_list = request.form.getlist('ids')
    
    if not action or not ids_list:
        flash("Aucune action ou aucun candidat sélectionné.", "warning")
        return redirect(url_for('liste_candidats'))
        
    try:
        if action == 'delete':
            candidats_data = load_json_file(config.CANDIDATS_FILE)
            candidats = candidats_data.get('candidats', [])
            
            # Fonction robuste pour déplacer/archiver même si Windows bloque
            def force_move(src, dst):
                if not os.path.exists(src): return
                try:
                    shutil.move(src, dst)
                except PermissionError:
                    import stat
                    if os.path.isdir(src):
                        shutil.copytree(src, dst, dirs_exist_ok=True)
                        shutil.rmtree(src, onerror=lambda f, p, e: (os.chmod(p, stat.S_IWRITE), f(p)))
                    else:
                        shutil.copy2(src, dst)
                        os.chmod(src, stat.S_IWRITE)
                        os.remove(src)

            count = 0
            for id_candidat in ids_list:
                candidat = next((c for c in candidats if c.get('id_candidat') == id_candidat), None)
                if not candidat: continue
                
                archive_timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                archive_dir = config.ARCHIVES_CANDIDATS_DIR / f"{id_candidat}_{archive_timestamp}"
                archive_dir.mkdir(parents=True, exist_ok=True)

                candidat_dir = config.CANDIDATS_DIR / id_candidat
                archived_candidat_dir = archive_dir / 'candidat_data'
                force_move(str(candidat_dir), str(archived_candidat_dir))

                pdf_file = config.EXPORT_DIR / f"programme_{id_candidat}.pdf"
                archived_pdf_file = archive_dir / pdf_file.name
                force_move(str(pdf_file), str(archived_pdf_file))

                excel_file = config.EXPORT_DIR / f"programme_{id_candidat}.xlsx"
                if excel_file.exists():
                    force_move(str(excel_file), str(archive_dir / excel_file.name))

                candidats = [c for c in candidats if c.get('id_candidat') != id_candidat]
                count += 1
                
            candidats_data['candidats'] = candidats
            save_json_file(config.CANDIDATS_FILE, candidats_data)
            flash(f"{count} candidat(s) archivé(s)/supprimé(s) avec succès.", "success")
            
        elif action == 'analyse':
            from data.analyse_besoins import charger_reponses, analyser_reponses, sauvegarder_programme
            candidats_data = load_json_file(config.CANDIDATS_FILE)
            count = 0
            for id_candidat in ids_list:
                try:
                    candidat = next((c for c in candidats_data.get('candidats', []) if c.get('id_candidat') == id_candidat), None)
                    if not candidat: continue
                    reponses = charger_reponses(id_candidat, str(config.BASE_DIR))
                    resultats = analyser_reponses(reponses, candidat.get('id_questionnaire'), str(config.BASE_DIR))
                    sauvegarder_programme(id_candidat, resultats, str(config.BASE_DIR))
                    
                    candidat['programme_genere'] = True
                    count += 1
                except Exception as e:
                    app.logger.error(f"Erreur génération programme {id_candidat}: {e}")
                    
            save_json_file(config.CANDIDATS_FILE, candidats_data)
            flash(f"{count} programme(s) généré(s) avec succès.", "success")
            
        elif action in ['pdf', 'excel']:
            import io
            import zipfile
            from src.pdf_generator import generer_pdf_programme
            from src.excel_generator import generer_excel_programme
            
            candidats_data = load_json_file(config.CANDIDATS_FILE)
            candidats = candidats_data.get('candidats', [])
            
            memory_file = io.BytesIO()
            with zipfile.ZipFile(memory_file, 'w', zipfile.DEFLATED) as zf:
                for id_candidat in ids_list:
                    candidat = next((c for c in candidats if c.get('id_candidat') == id_candidat), None)
                    if not candidat: continue
                    nom_fichier_base = f"programme_{candidat.get('nom', 'candidat')}_{candidat.get('prenom', '')}_{id_candidat}"
                    
                    try:
                        if action == 'pdf':
                            programme = analyser_candidat.charger_programme(id_candidat, str(config.BASE_DIR))
                            if programme:
                                pdf_path = config.EXPORT_DIR / f"tmp_{id_candidat}.pdf"
                                generer_pdf_programme(programme, candidat, pdf_path)
                                zf.write(pdf_path, arcname=f"{nom_fichier_base}.pdf")
                        elif action == 'excel':
                            excel_path = generer_excel_programme(id_candidat, str(config.BASE_DIR))
                            if excel_path and os.path.exists(excel_path):
                                zf.write(excel_path, arcname=f"{nom_fichier_base}.xlsx")
                    except Exception as e:
                        app.logger.error(f"Erreur export {action} {id_candidat}: {e}")
                        
            memory_file.seek(0)
            return send_file(
                memory_file,
                mimetype="application/zip",
                as_attachment=True,
                download_name=f"export_{action}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
            )
            
    except Exception as e:
        app.logger.error(f"Erreur action batch {action}: {e}")
        flash(f"Une erreur est survenue : {str(e)}", "danger")
        
    return redirect(url_for('liste_candidats'))


# ==================== API ====================

@app.route('/api/candidats')
def api_candidats():
    """API - Liste des candidats (index uniquement)"""
    try:
        candidats_data = load_json_file(config.CANDIDATS_FILE)
        return jsonify(candidats_data)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/candidats/<id_candidat>')
def api_candidat(id_candidat):
    """API - Détail d'un candidat avec ses réponses"""
    try:
        candidats_data = load_json_file(config.CANDIDATS_FILE)
        candidat = next((c for c in candidats_data['candidats'] if c['id_candidat'] == id_candidat), None)
        if not candidat:
            return jsonify({'error': 'Candidat non trouvé'}), 404

        # Enrichir avec les réponses du dossier
        questionnaire = charger_reponses_candidat(id_candidat)
        if questionnaire:
            candidat['reponses'] = questionnaire.get('reponses', [])

        return jsonify(candidat)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/programmes')
def api_programmes():
    """API - Liste des programmes disponibles"""
    try:
        programmes_data = load_json_file(config.PROGRAMMES_FILE)
        return jsonify(programmes_data)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/questionnaires')
def api_questionnaires():
    """API - Liste des questionnaires disponibles"""
    try:
        questions_data = load_json_file(config.QUESTIONS_FILE)
        # Retourner un résumé (sans toutes les questions)
        summary = []
        for q in questions_data.get('questionnaires', []):
            summary.append({
                'id_questionnaire': q['id_questionnaire'],
                'id_programme': q.get('id_programme'),
                'titre': q['titre'],
                'nb_questions': len(q.get('questions', []))
            })
        return jsonify({'questionnaires': summary})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/candidats/template-csv/<id_questionnaire>')
def telecharger_template_csv(id_questionnaire):
    """Génère et télécharge un Excel stylé questionnaire type pour candidats"""
    try:
        from src.excel_generator import generer_excel_questionnaire
        chemin = generer_excel_questionnaire(id_questionnaire, config.BASE_DIR)
        safe_id = id_questionnaire.lower().replace('quest_', '')
        filename = f"questionnaire_type_{safe_id}.xlsx"
        return send_file(
            chemin,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except ValueError as e:
        flash(str(e), 'error')
        return redirect(url_for('import_candidat'))
    except Exception as e:
        flash(f"Erreur lors de la génération du questionnaire : {str(e)}", 'error')
        return redirect(url_for('import_candidat'))


@app.route('/questionnaire/<id_questionnaire>')
def questionnaire_page(id_questionnaire):
    """Affiche le formulaire de questionnaire pour un candidat"""
    try:
        questions_data = load_json_file(config.QUESTIONS_FILE)
        import copy, re
        
        if id_questionnaire == 'QUEST_EXCEL':
            # Fusion des 3 questionnaires Excel (INIT/INTER/PERF) avec extraction de id_domaine et id_niveau
            domain_names = {
                'D1': 'Environnement / Méthodes',
                'D2': 'Calculs & Formules',
                'D3': 'Mise en forme',
                'D4': 'Gestion des données'
            }
            excel_quest_ids = ['QUEST_EXCEL_INIT', 'QUEST_EXCEL_INTER', 'QUEST_EXCEL_PERF']
            all_excel_questions = []
            seen_ids = set()
            for qid in excel_quest_ids:
                qsrc = next((q for q in questions_data['questionnaires'] if q['id_questionnaire'] == qid), None)
                if not qsrc:
                    continue
                for q in qsrc.get('questions', []):
                    if q['id_question'] in seen_ids:
                        continue
                    seen_ids.add(q['id_question'])
                    q_copy = copy.deepcopy(q)
                    # Extraire id_domaine et id_niveau depuis [DX-NX-CX]
                    m = re.match(r'\[?(D\d+)-(N\d+)-', q_copy['id_question'])
                    if m:
                        q_copy['id_domaine'] = m.group(1)
                        q_copy['id_niveau'] = m.group(2)
                    else:
                        q_copy.setdefault('id_domaine', 'D1')
                        q_copy.setdefault('id_niveau', 'N1')
                    all_excel_questions.append(q_copy)

            questionnaire = {
                'id_questionnaire': 'QUEST_EXCEL',
                'id_programme': 'PROG_EXCEL_INTER',
                'titre': 'Positionnement Excel - Tosa',
                'questions': all_excel_questions
            }
        else:
            questionnaire = next((q for q in questions_data['questionnaires'] if q['id_questionnaire'] == id_questionnaire), None)
            if questionnaire:
                questionnaire = copy.deepcopy(questionnaire)
                # Assigner id_niveau et id_domaine si manquants
                for q in questionnaire.get('questions', []):
                    m = re.match(r'\[?(D\d+)-(N\d+)-', q['id_question'])
                    if m:
                        q.setdefault('id_domaine', m.group(1))
                        q.setdefault('id_niveau', m.group(2))
        
        if not questionnaire:
            flash(f"Questionnaire {id_questionnaire} non trouvé", 'error')
            return redirect(url_for('dashboard'))
            
        programmes_data = load_json_file(config.PROGRAMMES_FILE)
        id_programme = questionnaire.get('id_programme', 'PROG_WORD')
        programme = next((p for p in programmes_data['programmes'] if p['id_programme'] == id_programme), None)
        
        # Grouper les questions par blocs (pour l'affichage séquentiel)
        questions_by_bloc = []
        all_questions = {q['id_question']: q for q in questionnaire.get('questions', [])}
        
        # Cas spécial pour Excel : Groupement par domaines TOSA
        if id_questionnaire == 'QUEST_EXCEL':
            domain_names = {
                'D1': 'Environnement / Méthodes',
                'D2': 'Calculs & Formules',
                'D3': 'Mise en forme',
                'D4': 'Gestion des données'
            }
            domaines_dict = {}
            for q in questionnaire.get('questions', []):
                dom = q.get('id_domaine', 'D1')
                if dom not in domaines_dict:
                    domaines_dict[dom] = {
                        'id_domaine': dom,
                        'nom_domaine': domain_names.get(dom, dom),
                        'questions': []
                    }
                domaines_dict[dom]['questions'].append(q)
            
            # Trier les domaines (D1, D2, D3, D4)
            sorted_domain_ids = sorted(domaines_dict.keys())
            sorted_domaines = [domaines_dict[d] for d in sorted_domain_ids]
            
            # On crée un seul bloc "Évaluation" qui contiendra les domaines
            questions_by_bloc.append({
                'id_bloc': 'EVAL',
                'nom_bloc': 'Évaluation des compétences',
                'domaines': sorted_domaines,
                'is_excel': True
            })
        elif programme and 'blocs_objectifs' in programme:
            for bloc in programme['blocs_objectifs']:
                bloc_questions = []
                for t_id in bloc.get('themes', []):
                    # Trouver toutes les questions de ce thème
                    for q_id, q_data in all_questions.items():
                        if q_data.get('id_theme') == t_id:
                            bloc_questions.append(q_data)
                
                if bloc_questions:
                    questions_by_bloc.append({
                        'id_bloc': bloc['id'],
                        'nom_bloc': bloc['nom'],
                        'questions': bloc_questions
                    })
        else:
            # Fallback: un seul bloc avec toutes les questions
            questions_by_bloc.append({
                'id_bloc': 'ALL',
                'nom_bloc': 'Questions',
                'questions': questionnaire.get('questions', [])
            })
        
        return render_template('candidat/formulaire.html', 
                             questionnaire=questionnaire,
                             programmes=programmes_data['programmes'],
                             questions_by_bloc=questions_by_bloc)
    except Exception as e:
        flash(f"Erreur: {str(e)}", 'error')
        return redirect(url_for('dashboard'))


@app.route('/questionnaire/submit', methods=['POST'])
def submit_questionnaire():
    """Traite la soumission du questionnaire et crée le candidat"""
    try:
        id_questionnaire = request.form.get('id_questionnaire')
        nom = request.form.get('nom', '').strip().upper()
        prenom = request.form.get('prenom', '').strip()
        id_programme = request.form.get('id_programme')
        self_level = request.form.get('self_level', 'debutant')
        
        if not nom or not prenom:
            flash("Le nom et le prénom sont obligatoires", 'error')
            return redirect(request.referrer)
            
        # Création de l'ID candidat (slug simple)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_nom = "".join(x for x in nom if x.isalnum())
        safe_prenom = "".join(x for x in prenom if x.isalnum())
        id_candidat = f"CAND_{safe_nom}_{safe_prenom}_{timestamp}"
        
        # Récupérer les réponses
        reponses = []
        
        # Ajouter une trace de profilage si Novice
        if self_level == 'novice':
            reponses.append({
                'id_question': 'INFO_NOVICE',
                'texte': "Auto-déclaration du niveau",
                'acquisition': 'Novice',
                'besoin': 'Oui (Débutant complet)'
            })
            
        for key, value in request.form.items():
            if key.startswith('acq_'):
                id_q = key.replace('acq_', '')
                acq = value
                bes = request.form.get(f'bes_{id_q}')
                reponses.append({
                    'id_question': id_q,
                    'acquisition': acq,
                    'besoin': bes
                })
        
        # Sauvegarder questionnaire.json dans le dossier candidat
        candidat_dir = config.CANDIDATS_DIR / id_candidat
        candidat_dir.mkdir(parents=True, exist_ok=True)
        
        quest_data = {
            "id_candidat": id_candidat,
            "id_questionnaire": id_questionnaire,
            "self_level": self_level,
            "date_reponse": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "reponses": reponses
        }
        
        save_json_file(candidat_dir / 'questionnaire.json', quest_data)
        
        # Mettre à jour l'index des candidats
        candidats_data = load_json_file(config.CANDIDATS_FILE)
        candidats_data['candidats'].append({
            "id_candidat": id_candidat,
            "nom": nom,
            "prenom": prenom,
            "self_level": self_level,
            "id_questionnaire": id_questionnaire,
            "id_programme": id_programme,
            "date_reponse": quest_data["date_reponse"]
        })
        save_json_file(config.CANDIDATS_FILE, candidats_data)
        
        # Lancer l'analyse
        generer_programme(id_candidat, candidat_dir / 'programme_perso.json')
        
        flash(f"Positionnement terminé pour {prenom} {nom}!", 'success')
        return redirect(url_for('detail_candidat', id_candidat=id_candidat))
        
    except Exception as e:
        flash(f"Erreur lors de la soumission: {str(e)}", 'error')
        return redirect(url_for('dashboard'))


# ==================== ERROR HANDLERS ====================

@app.errorhandler(404)
def page_not_found(e):
    return render_template('base.html', content="<div class='alert alert-error'>Page non trouvée</div>"), 404


@app.errorhandler(500)
def server_error(e):
    return render_template('base.html', content="<div class='alert alert-error'>Erreur serveur</div>"), 500


if __name__ == '__main__':
    print("=" * 50)
    print("  eS@deq - Plateforme de Positionnement")
    print("  http://localhost:5000")
    print("=" * 50)
    app.run(debug=config.DEBUG, host='0.0.0.0', port=5000)
